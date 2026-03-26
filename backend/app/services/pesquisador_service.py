"""
Servico do Pesquisador - Painel Cientifico

Funcionalidades:
  - Obter info dos modelos dos microsservicos
  - Gerar templates Excel para download
  - Validar arquivos de upload
  - Enviar dados para retreino dos microsservicos
  - Registrar e listar historico de uploads
"""
import json
import unicodedata
import logging
from io import BytesIO
from typing import List, Optional
from uuid import UUID
from datetime import datetime

import httpx
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import desc

from ..config import MICROSSERVICO_OLHO_PAVAO_URL, MICROSSERVICO_ANTRACNOSE_URL
from ..models.upload import UploadDados

logger = logging.getLogger(__name__)

_MICROSSERVICO_URLS = {
    "olho-pavao": MICROSSERVICO_OLHO_PAVAO_URL,
    "antracnose": MICROSSERVICO_ANTRACNOSE_URL,
}


# ============================================================
# OBTER INFO DOS MODELOS
# ============================================================

async def obter_info_modelos() -> list:
    """Busca info dos modelos de ambos os microsservicos."""
    resultados = []
    for doenca_id, base_url in _MICROSSERVICO_URLS.items():
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(f"{base_url}/modelo/info", timeout=10.0)
                response.raise_for_status()
                data = response.json()
                data['doenca_id'] = doenca_id
                resultados.append(data)
        except Exception as e:
            logger.warning(f"Microsservico {doenca_id} indisponivel: {e}")
            resultados.append({
                'doenca_id': doenca_id,
                'modelo': 'Indisponivel',
                'accuracy': 0,
                'f1_score': 0,
                'total_amostras_treino': 0,
                'anos_treino': [],
                'features_utilizadas': [],
                'thresholds': {},
            })
    return resultados


async def obter_info_modelo_unico(doenca_id: str) -> Optional[dict]:
    """Busca info de um modelo especifico."""
    base_url = _MICROSSERVICO_URLS.get(doenca_id)
    if not base_url:
        return None
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{base_url}/modelo/info", timeout=10.0)
            response.raise_for_status()
            data = response.json()
            data['doenca_id'] = doenca_id
            return data
    except Exception:
        return None


# ============================================================
# GERAR TEMPLATES EXCEL
# ============================================================

def _estilizar_header(ws, row, num_cols):
    """Aplica estilo ao cabecalho do Excel."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F7942", end_color="4F7942", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def gerar_template(tipo: str) -> BytesIO:
    """Gera template Excel com cabecalhos corretos e linha de exemplo."""
    wb = Workbook()
    ws = wb.active

    if tipo == "olho-pavao":
        ws.title = "Olho de Pavao"
        # Linha 1: titulo (porque pipeline usa header=1)
        ws.append(["Dados de campo - Olho de Pavao (preencha a partir da linha 3)"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=12, color="4F7942")
        # Linha 2: cabecalhos
        headers = ["data", "repeticao", "arvore", "folha", "visiveis", "visiveis + latentes"]
        ws.append(headers)
        _estilizar_header(ws, 2, len(headers))
        # Linha 3: exemplo
        ws.append(["2026-01-15", 1, 1, 1, 0, 0])
        ws.append(["2026-01-15", 1, 1, 2, 1, 2])
        # Ajustar largura
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18

    elif tipo == "antracnose":
        ws.title = "Antracnose"
        ws.append(["Dados de campo - Antracnose (preencha a partir da linha 3)"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=12, color="8B5CF6")
        headers = ["data", "olival/parcela", "arvore", "azeitona", "severidade", "incidencia"]
        ws.append(headers)
        _estilizar_header(ws, 2, len(headers))
        ws.append(["2026-10-15", "Parcela1", 1, 1, 0, 0])
        ws.append(["2026-10-15", "Parcela1", 1, 2, 1.5, 1])
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18

    elif tipo == "clima":
        ws.title = "Clima"
        # Clima nao usa header=1, cabecalhos na linha 1
        headers = ["ANO", "MES", "DIA", "T_MED", "T_MAX", "T_MIN", "HR_MED", "FF_MED", "PR_QTD"]
        ws.append(headers)
        _estilizar_header(ws, 1, len(headers))
        ws.append([2026, 1, 15, 12.5, 17.0, 8.0, 75.0, 2.5, 1.2])
        ws.append([2026, 1, 16, 13.0, 18.0, 9.0, 78.0, 1.8, 0.5])
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col_letter].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# VALIDACAO DE ARQUIVOS
# ============================================================

def _normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nomes das colunas: strip accents, lowercase, trim."""
    df.columns = [
        unicodedata.normalize('NFKD', str(c)).encode('ascii', 'ignore').decode().strip().lower()
        for c in df.columns
    ]
    return df


def validar_arquivo_doenca(df: pd.DataFrame, doenca_id: str) -> List[str]:
    """Valida colunas e dados do arquivo de doenca."""
    erros = []
    df = _normalizar_colunas(df)

    if doenca_id == "olho-pavao":
        required = ['data', 'repeticao', 'arvore', 'folha', 'visiveis', 'visiveis + latentes']
    else:
        required = ['data', 'olival/parcela', 'arvore', 'azeitona', 'severidade', 'incidencia']

    # Verificar colunas presentes
    colunas_existentes = list(df.columns)
    for col in required:
        if col not in colunas_existentes:
            erros.append(f"Coluna obrigatoria ausente: '{col}'")

    if erros:
        erros.append(f"Colunas encontradas: {colunas_existentes}")
        return erros

    # Verificar linhas
    if len(df) == 0:
        erros.append("Ficheiro sem dados (0 linhas)")
        return erros

    # Verificar valores nulos em TODAS as colunas
    for col in df.columns:
        nulos = df[col].isna().sum()
        if nulos > 0:
            linhas_nulas = df.index[df[col].isna()].tolist()[:5]
            erros.append(f"Coluna '{col}' tem {nulos} valor(es) nulo(s) (linhas: {linhas_nulas})")

    # Verificar strings vazias em TODAS as colunas de texto
    for col in df.columns:
        if df[col].dtype == object:
            vazios_mask = df[col].astype(str).str.strip() == ''
            vazios = vazios_mask.sum()
            if vazios > 0:
                linhas_vazias = df.index[vazios_mask].tolist()[:5]
                erros.append(f"Coluna '{col}' tem {vazios} valor(es) vazio(s) (linhas: {linhas_vazias})")

    if erros:
        return erros

    # Verificar data parseavel
    try:
        pd.to_datetime(df['data'])
    except Exception:
        erros.append("Coluna 'data' contem valores que nao sao datas validas")

    # Verificar colunas numericas
    if doenca_id == "olho-pavao":
        for col in ['visiveis', 'visiveis + latentes']:
            if col in df.columns:
                non_numeric = pd.to_numeric(df[col], errors='coerce').isna().sum()
                if non_numeric > 0:
                    erros.append(f"Coluna '{col}' tem {non_numeric} valores nao numericos")
    else:
        for col in ['severidade', 'incidencia']:
            if col in df.columns:
                non_numeric = pd.to_numeric(df[col], errors='coerce').isna().sum()
                if non_numeric > 0:
                    erros.append(f"Coluna '{col}' tem {non_numeric} valores nao numericos")

    return erros


def validar_arquivo_clima(df: pd.DataFrame) -> List[str]:
    """Valida colunas e dados do arquivo de clima."""
    erros = []

    # Clima usa nomes uppercase
    df.columns = [str(c).strip().upper() for c in df.columns]

    required = ['ANO', 'MES', 'DIA', 'T_MED', 'T_MAX', 'T_MIN', 'HR_MED', 'FF_MED', 'PR_QTD']
    colunas_existentes = list(df.columns)
    for col in required:
        if col not in colunas_existentes:
            erros.append(f"Coluna obrigatoria ausente: '{col}'")

    if erros:
        erros.append(f"Colunas encontradas: {colunas_existentes}")
        return erros

    if len(df) == 0:
        erros.append("Ficheiro sem dados (0 linhas)")
        return erros

    # Verificar valores nulos em TODAS as colunas
    for col in df.columns:
        nulos = df[col].isna().sum()
        if nulos > 0:
            linhas_nulas = df.index[df[col].isna()].tolist()[:5]
            erros.append(f"Coluna '{col}' tem {nulos} valor(es) nulo(s) (linhas: {linhas_nulas})")

    # Verificar strings vazias em TODAS as colunas de texto
    for col in df.columns:
        if df[col].dtype == object:
            vazios_mask = df[col].astype(str).str.strip() == ''
            vazios = vazios_mask.sum()
            if vazios > 0:
                linhas_vazias = df.index[vazios_mask].tolist()[:5]
                erros.append(f"Coluna '{col}' tem {vazios} valor(es) vazio(s) (linhas: {linhas_vazias})")

    if erros:
        return erros

    # Verificar ANO, MES, DIA sao inteiros
    for col in ['ANO', 'MES', 'DIA']:
        non_numeric = pd.to_numeric(df[col], errors='coerce').isna().sum()
        if non_numeric > 0:
            erros.append(f"Coluna '{col}' tem {non_numeric} valores nao numericos")

    # Verificar colunas climaticas sao numericas
    for col in ['T_MED', 'T_MAX', 'T_MIN', 'HR_MED', 'FF_MED', 'PR_QTD']:
        non_numeric = pd.to_numeric(df[col], errors='coerce').isna().sum()
        if non_numeric > 0:
            erros.append(f"Coluna '{col}' tem {non_numeric} valores nao numericos")

    return erros


# ============================================================
# ENVIAR RETREINO AO MICROSSERVICO
# ============================================================

def inserir_dados_upload(db: Session, doenca_id: str, df_doenca: pd.DataFrame, df_clima: pd.DataFrame):
    """Insere dados de upload nas tabelas de treino do banco."""
    from ..models.dados_treino import DadosOlhoPavao, DadosAntracnose, DadosClima

    # Normalizar colunas da doenca
    df_doenca = _normalizar_colunas(df_doenca)

    if doenca_id == "olho-pavao":
        registos = []
        for _, row in df_doenca.iterrows():
            registos.append(DadosOlhoPavao(
                data=pd.to_datetime(row.get('data')),
                repeticao=int(row['repeticao']) if pd.notna(row.get('repeticao')) else None,
                arvore=int(row['arvore']) if pd.notna(row.get('arvore')) else None,
                folha=int(row['folha']) if pd.notna(row.get('folha')) else None,
                visiveis=int(row['visiveis']) if pd.notna(row.get('visiveis')) else None,
                visiveis_latentes=int(row['visiveis + latentes']) if pd.notna(row.get('visiveis + latentes')) else None,
            ))
        db.bulk_save_objects(registos)
    else:
        registos = []
        for _, row in df_doenca.iterrows():
            registos.append(DadosAntracnose(
                data=pd.to_datetime(row.get('data')),
                olival_parcela=str(row['olival/parcela']) if pd.notna(row.get('olival/parcela')) else None,
                arvore=int(row['arvore']) if pd.notna(row.get('arvore')) else None,
                azeitona=int(row['azeitona']) if pd.notna(row.get('azeitona')) else None,
                severidade=float(row['severidade']) if pd.notna(row.get('severidade')) else None,
                incidencia=float(row['incidencia']) if pd.notna(row.get('incidencia')) else None,
            ))
        db.bulk_save_objects(registos)

    # Inserir dados de clima
    df_clima.columns = [str(c).strip().upper() for c in df_clima.columns]
    registos_clima = []
    for _, row in df_clima.iterrows():
        registos_clima.append(DadosClima(
            ano=int(row['ANO']),
            mes=int(row['MES']),
            dia=int(row['DIA']),
            t_med=float(row['T_MED']) if pd.notna(row.get('T_MED')) else None,
            t_max=float(row['T_MAX']) if pd.notna(row.get('T_MAX')) else None,
            t_min=float(row['T_MIN']) if pd.notna(row.get('T_MIN')) else None,
            hr_med=float(row['HR_MED']) if pd.notna(row.get('HR_MED')) else None,
            ff_med=float(row['FF_MED']) if pd.notna(row.get('FF_MED')) else None,
            pr_qtd=float(row['PR_QTD']) if pd.notna(row.get('PR_QTD')) else None,
        ))
    db.bulk_save_objects(registos_clima)
    db.commit()


async def enviar_retreino(doenca_id: str) -> dict:
    """Envia pedido de retreino ao microsservico (dados ja estao no banco)."""
    base_url = _MICROSSERVICO_URLS.get(doenca_id)
    if not base_url:
        raise Exception(f"Doenca desconhecida: {doenca_id}")

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{base_url}/modelo/retreinar",
                json={},
                timeout=120.0,
            )
            response.raise_for_status()
            return response.json()
    except httpx.ConnectError:
        raise Exception(f"Microsservico de {doenca_id} indisponivel.")
    except httpx.TimeoutException:
        raise Exception(f"Retreino demorou demais (timeout 120s).")
    except httpx.HTTPStatusError as e:
        detail = e.response.text
        raise Exception(f"Erro ao retreinar: {detail}")


# ============================================================
# HISTORICO DE UPLOADS
# ============================================================

def registrar_upload(
    db: Session,
    usuario_id: UUID,
    doenca_id: str,
    amostras_doenca: int,
    amostras_clima: int,
    anos_dados: List[int],
    accuracy_antes: Optional[float],
    accuracy_depois: float,
    f1_antes: Optional[float],
    f1_depois: float,
    total_amostras_depois: int,
) -> UploadDados:
    """Registra upload no historico."""
    upload = UploadDados(
        usuario_id=usuario_id,
        doenca_id=doenca_id,
        amostras_doenca=amostras_doenca,
        amostras_clima=amostras_clima,
        anos_dados=json.dumps(anos_dados),
        accuracy_antes=accuracy_antes,
        accuracy_depois=accuracy_depois,
        f1_antes=f1_antes,
        f1_depois=f1_depois,
        total_amostras_depois=total_amostras_depois,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)
    return upload


def listar_uploads(db: Session, pagina: int = 1, tamanho: int = 10) -> dict:
    """Lista historico de uploads com nome do usuario e paginacao."""
    from ..models.usuario import Usuario
    from sqlalchemy import func

    query = (
        db.query(UploadDados, Usuario.nome)
        .join(Usuario, UploadDados.usuario_id == Usuario.id)
        .order_by(desc(UploadDados.criado_em))
    )

    total = db.query(func.count(UploadDados.id)).scalar()
    total_paginas = max(1, (total + tamanho - 1) // tamanho)
    pagina = max(1, min(pagina, total_paginas))

    uploads = query.offset((pagina - 1) * tamanho).limit(tamanho).all()

    resultado = []
    for upload, nome in uploads:
        resultado.append({
            'id': str(upload.id),
            'doenca_id': upload.doenca_id,
            'usuario_nome': nome,
            'data_upload': upload.criado_em.isoformat() if upload.criado_em else None,
            'amostras_doenca': upload.amostras_doenca,
            'amostras_clima': upload.amostras_clima,
            'anos_dados': json.loads(upload.anos_dados) if upload.anos_dados else [],
            'accuracy_antes': float(upload.accuracy_antes) if upload.accuracy_antes else None,
            'accuracy_depois': float(upload.accuracy_depois),
            'f1_antes': float(upload.f1_antes) if upload.f1_antes else None,
            'f1_depois': float(upload.f1_depois),
            'total_amostras_depois': upload.total_amostras_depois,
        })

    return {
        'uploads': resultado,
        'pagina': pagina,
        'tamanho': tamanho,
        'total': total,
        'total_paginas': total_paginas,
    }
