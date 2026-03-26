"""
Script para criar ficheiros de teste de upload.
Gera ficheiros de erro (validacao) e ficheiros validos (5 e 20 dados).
"""
import pandas as pd
import openpyxl
import os
import random

random.seed(42)

output_dir = "C:/Users/maria/Downloads/testes_upload"
os.makedirs(output_dir, exist_ok=True)


# ============================================================
# HELPERS
# ============================================================

def criar_excel_doenca(filepath, titulo, colunas, dados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([titulo] + [""] * (len(colunas) - 1))
    ws.append(colunas)
    for row in dados:
        ws.append(row)
    wb.save(filepath)


def criar_excel_clima(filepath, dados):
    df = pd.DataFrame(dados, columns=["ANO", "MES", "DIA", "T_MED", "T_MAX", "T_MIN", "HR_MED", "FF_MED", "PR_QTD"])
    df.to_excel(filepath, index=False)


# ============================================================
# 1. ERROS - OLHO DE PAVAO
# ============================================================

TITULO_PAVAO = "Dados de campo - Olho de Pavao (preencha a partir da linha 3)"
COLS_PAVAO = ["data", "repeticao", "arvore", "folha", "visiveis", "visiveis + latentes"]

# 1a. Colunas faltando
criar_excel_doenca(
    f"{output_dir}/erro_pavao_colunas_faltando.xlsx",
    TITULO_PAVAO,
    ["data", "repeticao", "arvore"],
    [["2026-01-15", 1, 1]],
)

# 1b. Valores nulos
criar_excel_doenca(
    f"{output_dir}/erro_pavao_valores_nulos.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [
        ["2026-01-15", 1, 1, 1, 0, 0],
        ["2026-01-15", 1, 1, 2, None, 2],
        ["2026-01-15", None, 1, 3, 0, 0],
    ],
)

# 1c. Strings vazias
criar_excel_doenca(
    f"{output_dir}/erro_pavao_strings_vazias.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [
        ["2026-01-15", 1, 1, 1, 0, 0],
        ["", 1, 1, 2, 1, 2],
        ["2026-01-15", "", 1, 3, 0, 0],
    ],
)

# 1d. Datas invalidas
criar_excel_doenca(
    f"{output_dir}/erro_pavao_datas_invalidas.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [
        ["2026-01-15", 1, 1, 1, 0, 0],
        ["nao-e-data", 1, 1, 2, 1, 2],
        ["32/13/2026", 1, 1, 3, 0, 0],
    ],
)

# 1e. Valores nao numericos
criar_excel_doenca(
    f"{output_dir}/erro_pavao_nao_numerico.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [
        ["2026-01-15", 1, 1, 1, 0, 0],
        ["2026-01-15", "abc", 1, 2, 1, 2],
        ["2026-01-15", 1, 1, 3, "xyz", 0],
    ],
)

# 1f. Ficheiro vazio
criar_excel_doenca(
    f"{output_dir}/erro_pavao_vazio.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [],
)

print("Erros Olho de Pavao criados!")


# ============================================================
# 2. ERROS - ANTRACNOSE
# ============================================================

TITULO_ANTRAC = "Dados de campo - Antracnose (preencha a partir da linha 3)"
COLS_ANTRAC = ["data", "olival/parcela", "arvore", "azeitona", "severidade", "incidencia"]

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_colunas_faltando.xlsx",
    TITULO_ANTRAC,
    ["data", "olival/parcela"],
    [["2026-10-15", "Parcela1"]],
)

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_valores_nulos.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [
        ["2026-10-15", "Parcela1", 1, 1, 0, 0],
        ["2026-10-15", "Parcela1", 1, 2, None, 1],
        ["2026-10-15", None, 1, 3, 0, 0],
    ],
)

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_strings_vazias.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [
        ["2026-10-15", "Parcela1", 1, 1, 0, 0],
        ["", "Parcela1", 1, 2, 1.5, 1],
        ["2026-10-15", "", 1, 3, 0, 0],
    ],
)

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_datas_invalidas.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [
        ["2026-10-15", "Parcela1", 1, 1, 0, 0],
        ["texto-invalido", "Parcela1", 1, 2, 1.5, 1],
        ["99/99/9999", "Parcela1", 1, 3, 0, 0],
    ],
)

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_nao_numerico.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [
        ["2026-10-15", "Parcela1", 1, 1, 0, 0],
        ["2026-10-15", "Parcela1", "abc", 2, 1.5, 1],
        ["2026-10-15", "Parcela1", 1, 3, "xyz", 0],
    ],
)

criar_excel_doenca(
    f"{output_dir}/erro_antracnose_vazio.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [],
)

print("Erros Antracnose criados!")


# ============================================================
# 3. ERROS - CLIMA
# ============================================================

df = pd.DataFrame({"ANO": [2026], "MES": [1], "DIA": [15]})
df.to_excel(f"{output_dir}/erro_clima_colunas_faltando.xlsx", index=False)

criar_excel_clima(f"{output_dir}/erro_clima_valores_nulos.xlsx", [
    [2026, 1, 15, 12.5, 17, 8, 75, 2.5, 1.2],
    [2026, 1, 16, None, 18, 9, 78, 1.8, 0.5],
    [2026, 1, 17, 11.0, 16, None, 80, 2.0, 0.0],
])

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ANO", "MES", "DIA", "T_MED", "T_MAX", "T_MIN", "HR_MED", "FF_MED", "PR_QTD"])
ws.append([2026, 1, 15, 12.5, 17, 8, 75, 2.5, 1.2])
ws.append([2026, 1, 16, "", 18, 9, 78, 1.8, 0.5])
ws.append([2026, 1, 17, 11.0, 16, 7, "", 2.0, 0.0])
wb.save(f"{output_dir}/erro_clima_strings_vazias.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ANO", "MES", "DIA", "T_MED", "T_MAX", "T_MIN", "HR_MED", "FF_MED", "PR_QTD"])
ws.append([2026, 1, 15, 12.5, 17, 8, 75, 2.5, 1.2])
ws.append([2026, 1, 16, "abc", 18, 9, 78, 1.8, 0.5])
ws.append([2026, 1, 17, 11.0, "xyz", 7, 80, 2.0, 0.0])
wb.save(f"{output_dir}/erro_clima_nao_numerico.xlsx")

df = pd.DataFrame(columns=["ANO", "MES", "DIA", "T_MED", "T_MAX", "T_MIN", "HR_MED", "FF_MED", "PR_QTD"])
df.to_excel(f"{output_dir}/erro_clima_vazio.xlsx", index=False)

print("Erros Clima criados!")


# ============================================================
# 4. VALIDOS - OLHO DE PAVAO 5
# ============================================================

criar_excel_doenca(
    f"{output_dir}/valido_pavao_5.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    [
        ["2026-03-05", 1, 1, 1, 0, 0],
        ["2026-03-05", 1, 1, 2, 1, 2],
        ["2026-03-05", 1, 2, 1, 0, 1],
        ["2026-03-12", 1, 1, 1, 2, 3],
        ["2026-03-12", 1, 2, 1, 0, 0],
    ],
)

criar_excel_clima(f"{output_dir}/valido_clima_pavao_5.xlsx", [
    [2026, 3, 3, 11.2, 16, 6, 78, 2.1, 1.5],
    [2026, 3, 4, 12.0, 17, 7, 80, 1.8, 2.0],
    [2026, 3, 5, 10.5, 15, 5, 82, 2.3, 0.8],
    [2026, 3, 10, 13.0, 18, 8, 76, 1.5, 0.0],
    [2026, 3, 11, 14.2, 19, 9, 74, 1.2, 0.0],
    [2026, 3, 12, 12.8, 17, 8, 79, 2.0, 3.2],
    [2026, 3, 13, 11.5, 16, 7, 81, 2.5, 1.0],
])

print("Valido Olho de Pavao 5 criado!")


# ============================================================
# 5. VALIDOS - OLHO DE PAVAO 20
# ============================================================

dados_pavao_20 = []
for semana_offset in range(5):
    base_date = f"2026-03-{5 + semana_offset * 7:02d}"
    for rep in range(1, 3):
        for arv in range(1, 3):
            vis = random.randint(0, 3)
            vis_lat = vis + random.randint(0, 2)
            dados_pavao_20.append([base_date, rep, arv, len(dados_pavao_20) % 10 + 1, vis, vis_lat])

criar_excel_doenca(
    f"{output_dir}/valido_pavao_20.xlsx",
    TITULO_PAVAO,
    COLS_PAVAO,
    dados_pavao_20,
)

clima_pavao_20 = []
for d in range(1, 36):
    dia = ((d - 1) % 28) + 1
    mes = 3
    t_med = round(10 + random.uniform(-2, 5), 1)
    t_max = int(round(t_med + random.uniform(3, 8), 0))
    t_min = int(round(t_med - random.uniform(3, 6), 0))
    hr = int(round(70 + random.uniform(-5, 15), 0))
    ff = round(random.uniform(0.5, 4.0), 1)
    pr = round(max(0, random.uniform(-0.5, 5.0)), 1)
    clima_pavao_20.append([2026, mes, dia, t_med, t_max, t_min, hr, ff, pr])

criar_excel_clima(f"{output_dir}/valido_clima_pavao_20.xlsx", clima_pavao_20)

print("Valido Olho de Pavao 20 criado!")


# ============================================================
# 6. VALIDOS - ANTRACNOSE 5
# ============================================================

criar_excel_doenca(
    f"{output_dir}/valido_antracnose_5.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    [
        ["2026-10-15", "Parcela1", 1, 1, 0, 0],
        ["2026-10-15", "Parcela1", 1, 2, 1.5, 1],
        ["2026-10-15", "Parcela1", 2, 1, 0, 0],
        ["2026-10-22", "Parcela1", 1, 1, 2.0, 1],
        ["2026-10-22", "Parcela1", 2, 1, 0.5, 1],
    ],
)

criar_excel_clima(f"{output_dir}/valido_clima_antracnose_5.xlsx", [
    [2026, 10, 13, 18.5, 24, 13, 72, 1.5, 0.0],
    [2026, 10, 14, 19.2, 25, 14, 75, 1.2, 0.5],
    [2026, 10, 15, 17.8, 23, 12, 78, 2.0, 2.5],
    [2026, 10, 20, 20.0, 26, 15, 80, 1.0, 3.0],
    [2026, 10, 21, 21.5, 27, 16, 82, 0.8, 4.5],
    [2026, 10, 22, 19.8, 25, 14, 79, 1.5, 1.2],
    [2026, 10, 23, 18.0, 24, 13, 76, 2.2, 0.0],
])

print("Valido Antracnose 5 criado!")


# ============================================================
# 7. VALIDOS - ANTRACNOSE 20
# ============================================================

dados_antrac_20 = []
for semana_offset in range(5):
    base_date = f"2026-10-{8 + semana_offset * 7:02d}"
    for arv in range(1, 3):
        for az in range(1, 3):
            sev = round(random.uniform(0, 3), 1)
            inc = 1 if sev > 0.5 else 0
            dados_antrac_20.append([base_date, "Parcela1", arv, az, sev, inc])

criar_excel_doenca(
    f"{output_dir}/valido_antracnose_20.xlsx",
    TITULO_ANTRAC,
    COLS_ANTRAC,
    dados_antrac_20,
)

clima_antrac_20 = []
for d in range(1, 36):
    dia = 7 + ((d - 1) % 24)
    mes = 10
    t_med = round(17 + random.uniform(-3, 6), 1)
    t_max = int(round(t_med + random.uniform(4, 8), 0))
    t_min = int(round(t_med - random.uniform(3, 5), 0))
    hr = int(round(72 + random.uniform(-5, 15), 0))
    ff = round(random.uniform(0.5, 3.5), 1)
    pr = round(max(0, random.uniform(-1, 8.0)), 1)
    clima_antrac_20.append([2026, mes, dia, t_med, t_max, t_min, hr, ff, pr])

criar_excel_clima(f"{output_dir}/valido_clima_antracnose_20.xlsx", clima_antrac_20)

print("Valido Antracnose 20 criado!")


# ============================================================
# RESUMO
# ============================================================

print("\n" + "=" * 60)
print("TODOS OS FICHEIROS CRIADOS em:", output_dir)
print("=" * 60)
for f in sorted(os.listdir(output_dir)):
    size = os.path.getsize(os.path.join(output_dir, f))
    print(f"  {f} ({size} bytes)")
